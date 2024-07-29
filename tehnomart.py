import requests
from bs4 import BeautifulSoup
import json
import openpyxl
from datetime import datetime
import re

host = 'https://texnomart.uz'
base_url = host + '/ru/katalog/smartfony/'


def get_soup(url):
    html = requests.get(url).text       # получение HTML-кода страницы
    return BeautifulSoup(html, "html.parser")   # возврат объекта BeautifulSoup для парсинга


def get_analyst(base_url, category_name):
    result = {category_name: []}  # инициализация в виде словаря
    page = 1  # начальный номер страницы

    while True:
        """Цикл для обхода всех страниц пагинации"""
        url = f"{base_url}?page={page}"    # формирование URL для текущей страницы
        print(f"Парсинг страницы: {url}")  # Вывод текущей страницы
        soup = get_soup(url)    # получение HTML-кода и создание объекта BeautifulSoup
        product_box = soup.find('div', class_="products-box")    # поиск блока с продуктами

        # если блок с продуктами не найден
        if product_box is None:
            print(f"Нет блока с продуктами на странице {url}")  # сообщение об ошибке
            break        # выход из цикла

        items = product_box.find_all('div', class_='col-3')  # поиск товаров на странице

        # обход всех товаров на странице
        if not items:
            print(f"Нет товаров на странице {url}")   # поиск названия
            break

        for item in items:
            # извлечение названия модели
            full_title = item.find('div', class_='product-bottom__left').find('a').get_text(strip=True)
            match = re.search(r'(.*?)Количество камер:',
                              full_title)  # удаляем слово "Количество камер" и все что за ней
            title = match.group(1).strip() if match else full_title
            print(f"Модель: {title}")

            # извлечение и обработка цены
            price_text = item.find('div', class_='product-price__current').get_text(strip=True)
            price = float(price_text.replace('сум', '').replace(' ', '').replace(',', '.'))  # удаление слова "сум" и преобразование в дробное число
            print(f"Цена: {price}")

            # извлечение ссылки на детальную страницу товара
            link = host + item.find('a')['href']
            print(f"Ссылка: {link}")

            result[category_name].append({
                "Модель": title,
                "Цена": price,
                "Ссылка": link
            })

        page += 1

    # Сохранение данных
    save_data(result, category_name)
    return result


def save_data(data, category_name):
    # сохранение данных в JSON файл
    with open('tehnomart.json', mode='w', encoding='utf-8') as file:
        json.dump(data, file, indent=4, ensure_ascii=False)   # запись данных в файл

    # сохранение данных в Excel файл
    save_to_excel(data, category_name)


def save_to_excel(data, category_name):
    filename = 'tehnomart.xlsx'    # имя файла для сохранения
    today = datetime.today().strftime('%Y-%m-%d')   # текущая дата в формате год-месяц-день

    # загрузка или создание рабочей книги
    try:
        workbook = openpyxl.load_workbook(filename)  # попытка загрузить существующую книгу
    except FileNotFoundError:
        workbook = openpyxl.Workbook()    # создание новой книги
        workbook.active.title = category_name    # установка имени листа
    sheet = workbook[category_name]     # выбор листа по имени

    # добавление заголовков, если лист новый
    if sheet.max_row == 1 and sheet.max_column == 1:
        sheet.append(["Модель", "Цена", "Ссылка", "Дата"])  # запись заголовка таблицы

    # добавление новых данных
    for item in data[category_name]:
        sheet.append([item["Модель"], item["Цена"], item["Ссылка"], today])  # запись данных

    # расчёт и добавление средней цены
    prices = [item["Цена"] for item in data[category_name]]   # извлечение цен
    avg_price = sum(prices) / len(prices) if prices else 0    # расчет средней цены
    sheet.append(["Средняя цена", avg_price])       # заспись средней цены

    # сохранение рабочей книги
    workbook.save(filename)


if __name__ == "__main__":
    print(get_analyst(base_url=base_url, category_name='tehnomart'))
